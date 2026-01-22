# coding=utf-8
import pandas as pd
import glob
import sys
import os
import warnings
from copy import copy
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

print("环境初期化...")
strDatatime = datetime.now().strftime("%Y%m%d_%H%M%S")
if getattr(sys, 'frozen', False):
    # 如果是 PyInstaller 打包后的 exe
    BASE_DIR = os.path.dirname(sys.executable)
else:
    # 正常 python 运行
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
input_path  = os.path.join(BASE_DIR, "input")
output_path  = os.path.join(BASE_DIR, "output")
output_file = os.path.join(output_path, "报告查询合并结果_{0}.xlsx")

if not os.path.exists(input_path):
    print(f"输入目录 {input_path} 不存在，请重新确认")
    print("处理终止！")
    exit(1)

if not os.path.exists(output_path):
    print(f"创建输出目录 {output_path} ...")
    os.makedirs(output_path)

print("处理开始...")
# 找到所有 xlsx 文件并排序，以保证顺序稳定
files = sorted(glob.glob(os.path.join(input_path, "*.xlsx")))
if not files:
    print(f"输入目录 {input_path} 下没有找到任何 Excel 文件")
    print("处理终止！")
    exit(1)

all_dfs = []

# 用第一个文件的 A1 单元格作为“样式模板”
template_file = files[0]
# 第一个 Excel 的第一个 Sheet 名作为Sheet名
wb_tpl = load_workbook(template_file)
ws_tpl = wb_tpl.active
sheet_name = ws_tpl.title

for i, file in enumerate(files, 1):
    try:
        print(f"[{i}/{len(files)}] 正在处理：{os.path.basename(file)}")
        df = pd.read_excel(file, dtype=str)

        if df.empty:
            print(f"跳过空文件：{os.path.basename(file)}")
            continue

        # 在第一列插入“来源文件”列，方便追踪
        df.insert(0, "来源文件", os.path.basename(file))

        all_dfs.append(df)

    except Exception as e:
        print(f"读取失败: {os.path.basename(file)}, 错误: {e}")

if not all_dfs:
    print("所有 Excel 文件都是空的，没有可合并的数据")
    print("处理终止！")
    exit(1)

print("合并开始...")
# 合并（自动对齐列名，缺的补空）
output_file = output_file.format(strDatatime)
merged_df = pd.concat(all_dfs, ignore_index=True, sort=False)
merged_df.to_excel(output_file, index=False, sheet_name=sheet_name)

print("表头样式统一...")
# 读取模板文件，取 A1 的样式
wb_tpl = load_workbook(template_file)
ws_tpl = wb_tpl.active
cell_tpl = ws_tpl["A1"]

# 读取新生成的合并文件
wb_new = load_workbook(output_file)
ws_new = wb_new.active

# 把 A1 的样式复制到新文件第一行的所有单元格
for col in range(1, ws_new.max_column + 1):
    cell_new = ws_new.cell(row=1, column=col)

    if cell_tpl.has_style:
        cell_new.font = copy(cell_tpl.font)
        cell_new.border = copy(cell_tpl.border)
        cell_new.fill = copy(cell_tpl.fill)
        cell_new.number_format = copy(cell_tpl.number_format)
        cell_new.protection = copy(cell_tpl.protection)
        cell_new.alignment = copy(cell_tpl.alignment)

print("首行表头冻结...")
# 冻结首行（滚动时表头不动）
ws_new.freeze_panes = "A2"

print("首行表头填加筛选按钮...")
# 给表头加筛选按钮
ws_new.auto_filter.ref = ws_new.dimensions

print("列宽自动调整...")
for col in ws_new.columns:
    max_length = 0
    col_letter = get_column_letter(col[0].column)

    for cell in col:
        if cell.value:
            max_length = max(max_length, len(str(cell.value)))

    ws_new.column_dimensions[col_letter].width = min(max_length + 2, 50)

print("合并文件保存...")
wb_new.save(output_file)

print("输出文件：", output_file)
print("合并完成！！！")

if getattr(sys, 'frozen', False):
    print("\n处理完成，按任意键退出...")
    import msvcrt
    msvcrt.getch()
